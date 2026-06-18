from openpyxl.styles import PatternFill
from enum import StrEnum

class Warning_color(StrEnum):
    NEUTRAL = "D9D9D9" #світло-сірий
    ORDER_NOT_FOUND = "FF9999"  # світло-червоний
    MULTIPLE_RECORD = "A9D9A9"  # світло-зелений
    DATES_PROBLEM = "D9E2F3"  # світло-блакитний


def mark_row_wcolor(ws, row, fgColor="D9D9D9"):
    color_fill = PatternFill(
        fill_type="solid",
        fgColor=fgColor  # світло-сірий
    )
    for cell in ws[row]:
        cell.fill = color_fill

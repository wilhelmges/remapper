from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from datetime import datetime, date
from core import is_valid_date

def validate_cancels(ws):
    for row_num in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=2)
        if is_valid_date(cell.valur) and cell.font.striked:
            pass

def prepare_headers(ws):
    ws["B1"] = "дата"
    ws["C1"] = "номер"
    ws["Y1"]='comment'
    for col in range(4, 25):  # 3..10 включно
        cell = ws.cell(row=1, column=col)
        cell.value = str(cell.value).lower().strip()
        print(cell.value)
    ws["V1"] = "елтех"

def prepare_rows(ws):
    last_row = ws.max_row
    row = 3
    id=1
    while True:
        value_b = ws.cell(row=row, column=2).value
        text = str(value_b).strip()

        if row>last_row:
            return
        if not is_valid_date(value_b):# not text or text is None:
            row+=1
            continue


        ws.cell(row=row, column=1).value = id
        id += 1
        row+=1

def migrate_excel_sheme(ws:Worksheet):
    ws.insert_cols(4, amount=3)
    ws["C1"] = "operator"
    ws["D1"] = "order_id"
    ws["E1"] = "effects"
    ws["F1"] = "comment"

def prepare_excel(sourcefile):
    wb = load_workbook(sourcefile);  ws:Worksheet = wb["Sheet1"]
    prepare_headers(ws)
    wb.save(sourcefile) #;wb = load_workbook(sourcefile);  ws:Worksheet = wb["Sheet1"]

if __name__ == "__main__":
    prepare_excel()

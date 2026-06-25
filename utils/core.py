from datetime import datetime, date
import re
from typing import Optional
from decimal import Decimal, InvalidOperation
import math

import hashlib
from openpyxl.worksheet.worksheet import Worksheet
from datetime import date, datetime
from openpyxl.utils.datetime import from_excel


# sourcefile = "накази_втрати майна  А4007.xlsx"
# outputfile = "книга втрат електронний варіант.xlsx"

def is_valid_date(value):
    if value is None:
        return False

    if isinstance(value, (datetime, date)):
        return True

    if not isinstance(value, str):
        return False

    value = value.strip()
    if not value:
        return False

    formats = (
        "%d.%m.%Y",
        "%d.%m.%y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
    )

    for fmt in formats:
        try:
            datetime.strptime(value, fmt)
            return True
        except ValueError:
            pass

    return False

def cell_to_sqlite_date(cell) -> str | None:
    """
    Перетворює будь-яку Excel-ячейку в дату формату YYYY-MM-DD.

    Повертає:
        '2024-09-02' або None.
    """

    value = cell.value

    # порожня ячейка
    if value is None:
        return None

    # datetime
    if isinstance(value, datetime):
        return value.date().isoformat()

    # date
    if isinstance(value, date):
        return value.isoformat()

    # Excel serial date (число)
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return None

    # рядок
    if isinstance(value, str):

        value = value.strip()

        if not value:
            return None

        formats = (
            "%d.%m.%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%d-%m-%Y",
        )

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                pass

        return None

    return None

_NUMBER_RE = re.compile(r"(\d+)\s*$")
def get_order_from_comment(s) -> Optional[int]:
    if s is None:
        return None

    operation = str(s).strip()
    if not operation:
        return None

    operation = re.sub(r"\([^)]*\)", "", operation).strip()

    match = _NUMBER_RE.search(operation)
    if match:
        return int(match.group(1))

    return None

def cell_to_decimal(cell) -> Decimal | None:
    """
    Безпечно перетворює значення openpyxl.cell.Cell у Decimal.

    Повертає:
        Decimal - якщо значення коректне.
        None    - якщо значення відсутнє або некоректне.
    """
    value = cell.value

    if value is None:
        return None

    # Уже Decimal
    if isinstance(value, Decimal):
        return value

    # Цілі числа
    if isinstance(value, int):
        return Decimal(value)

    # float з Excel
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        # Через str(), щоб не тягнути двійкові похибки float
        return Decimal(str(value))

    # Рядок
    if isinstance(value, str):
        s = (
            value.strip()
            .replace("\xa0", "")   # нерозривний пробіл
            .replace(" ", "")      # звичайний пробіл
            .replace(",", ".")     # десяткова кома -> крапка
        )

        if not s:
            return None

        try:
            return Decimal(s)
        except InvalidOperation:
            return None

    # Інші типи
    return None

def calculate_md5(file_path) -> str:
    md5 = hashlib.md5()
    with open(file_path, "rb") as f:
        while chunk := f.read(1024 * 1024):
            md5.update(chunk)
    return md5.hexdigest()

def is_valid_order_row(ws, row):
    if ws.cell(row=row, column=2).value is None or ws.cell(row=row, column=3).value is None:
        return False
    textdate = ws.cell(row=row, column=2).value
    if not is_valid_date(textdate):
        print('cant get order or data', row, textdate)
        return False
    return True


def get_order_from_comment(s="(зміни в 2431)    3719"):
    operation = str(s).strip()
    if operation.isdigit():
        order_id = int(operation)
    else:
        order_id = re.sub(r'\([^)]*\)', '', operation).strip()
        order_id = order_id.split()[-1]
    return order_id

def remove_last_number(s="(зміни в 2431)                               3719"):
    if s is None:
        return None
    s=str(s)
    m = re.search(r'^(.*?)([+-]?\d+)\s*$', s)
    return m.group(1).rstrip() if m else s

def format_date_for_output(date_str):
    return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")

def delete_empty_bottom(ws:Worksheet):
    start_row = None
    empty_count = 0

    for row in range(ws.max_row, 0, -1):
        is_empty = all(
            cell.value in (None, '')
            for cell in ws[row]
        )

        if is_empty:
            start_row = row
            empty_count += 1
        else:
            break

    if empty_count > 0:
        print('deleting ', empty_count)
        ws.delete_rows(start_row, empty_count)
    else:
        print('no empty rows')

headers = ['рао', 'рао збб та р', 'зас ураж', 'бпла', 'ппо', 'нсо', 'реб', 'овт та мсп', 'реч', 'інж', 'зв', 'рхбз', 'ас', 'прод', 'мед', 'пмм', 'гео', 'кес', 'елтех', 'пожежна', 'метрол']
sheets = ['БпЛА', 'ОВТ', 'ЗВ', 'ЗББ', 'ЗУ', 'РЕЧ', 'НСО (БТ)', 'Ел-тех', 'ІС', 'ГЕО', 'прод', 'пмм', 'СВТ (АС)', 'мед', 'КЕС( СІ-ІЗ)', 'Метрологія']

department_to_sheet = {
    None: None,
    "рао": None,
    "рао збб та р": "ЗББ",
    "зас ураж": "ЗУ",
    "бпла": "БпЛА",
    "ппо": None,
    "нсо": "НСО (БТ)",
    "реб": None,
    "овт та мсп": "ОВТ",
    "реч": "РЕЧ",
    "інж": "ІС",
    "зв": "ЗВ",
    "рхбз": None,
    "ас": "СВТ (АС)",
    "прод": "прод",
    "мед": "мед",
    "пмм": "пмм",
    "гео": "ГЕО",
    "кес": "КЕС( СІ-ІЗ)",
    "елтех": "Ел-тех",
    "пожежна": None,
    "метрол": "Метрологія",
}

if __name__=='__main__':
    print(format_date_for_output('2026-03-17'))

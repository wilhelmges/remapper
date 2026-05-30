import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re

sourcefile = "source.xlsx"

from datetime import datetime, date

def is_valid_date(value):
    if value is None:
        return False

    if isinstance(value, (datetime, date)):
        return True

    if not isinstance(value, str):
        return False

    value = value.strip()
    if not value:
        return False

    formats = (
        "%d.%m.%Y",
        "%d.%m.%y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
    )

    for fmt in formats:
        try:
            datetime.strptime(value, fmt)
            return True
        except ValueError:
            pass

    return False

def validate_cancels(ws):
    for row_num in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=2)
        if is_valid_date(cell.valur) and cell.font.striked:
            pass

def prepare_headers(ws):
    ws["B1"] = "дата"
    ws["C1"] = "номер"
    ws["Y1"]='comment'
    for col in range(4, 25):  # 3..10 включно
        cell = ws.cell(row=1, column=col)
        cell.value = str(cell.value).lower().strip()
        print(cell.value)
    ws["V1"] = "елтех"

def remove_last_number(s="(зміни в 2431)                               3719"):
    if s is None:
        return None
    s=str(s)
    m = re.search(r'^(.*?)([+-]?\d+)\s*$', s)
    return m.group(1).rstrip() if m else s


def prepare_rows(ws):
    last_row = ws.max_row
    row = 3
    id=1
    while True:
        value_b = ws.cell(row=row, column=2).value
        text = str(value_b).strip()

        if row>last_row:
            return
        if not is_valid_date(value_b):# not text or text is None:
            row+=1
            continue


        ws.cell(row=row, column=1).value = id
        #operation = ws.cell(row=row, column=3).value
        # order_id = get_last_number(operation)
        effect = remove_last_number(operation)
        ws.cell(row=row, column=1).value = id
        # ws.cell(row=row, column=4).value = order_id
        # ws.cell(row=row, column=5).value = effect
        # ws.cell(row=row, column=6).value = ws.cell(row=row, column=28).value

        print(row, id, order_id, effect)
        id += 1
        row+=1

def migrate_excel_sheme(ws:Worksheet):
    ws.insert_cols(4, amount=3)
    ws["C1"] = "operator"
    ws["D1"] = "order_id"
    ws["E1"] = "effects"
    ws["F1"] = "comment"

def delete_empty_bottom(ws:Worksheet):
    start_row = None
    empty_count = 0

    for row in range(ws.max_row, 0, -1):
        is_empty = all(
            cell.value in (None, '')
            for cell in ws[row]
        )

        if is_empty:
            start_row = row
            empty_count += 1
        else:
            break

    if empty_count > 0:
        print('deleting ', empty_count)
        ws.delete_rows(start_row, empty_count)
    else:
        print('no empty rows')

def prepare_excel():
    shutil.copy2("накази_втрати майна  А4007.xlsx", sourcefile)
    wb = load_workbook(sourcefile);  ws:Worksheet = wb["Sheet1"]
    #print(ws['B9177'].font.strike); return
    prepare_headers(ws)

    prepare_rows(ws)
    wb.save(sourcefile) #;wb = load_workbook(sourcefile);  ws:Worksheet = wb["Sheet1"]


if __name__ == "__main__":
    prepare_excel()

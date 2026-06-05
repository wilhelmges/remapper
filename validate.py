import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import sqlite3
import re
from core import is_valid_date
import datetime
from styles import mark_neutral,worning_color

from core import headers, department_to_sheet, outputfile
import re

from core import get_order_from_comment


def get_order(s= "№106 (ОД) 08.01.2026"):
    match = re.search(r"\d+", s)
    num = int(match.group()) if match else None
    return num

if __name__=="__main__":
    year = 2026
    wb = load_workbook(outputfile, data_only=True)
    for ws in wb.worksheets:

        sheet_name = ws.title
        print(sheet_name)

        row = 8
        last_row = ws.max_row #14#
        for row in range(8, last_row):
            date = ws.cell(row=row, column=1).value
            if not is_valid_date(date):
                continue
            if type(date) == datetime.datetime:
                date = date.date()
            if type(date) == str:
                date = datetime.datetime.strptime(date, "%d.%m.%Y").date()
            order_id = get_order(ws.cell(row=row, column=2).value)
            d = (ws.cell(row=row, column=2).value.split()[-1])
            try:
                if len(d)==10:
                    date2 = datetime.datetime.strptime(d, "%d.%m.%Y").date()
                elif len(d)==8:
                    date2 = datetime.datetime.strptime(d, "%d.%m.%y").date()
                else:
                    date2 = None; print(" cant convert date: ", date, sheet_name, row, repr(d))
                    mark_neutral(ws, row, worning_color.DATES_PROBLEM)
            except Exception as e:
                print(" error while converting to date: ",date, sheet_name, row, repr(d))
                mark_neutral(ws, row, worning_color.DATES_PROBLEM)
                date2 = date

            if date!=date2:
                print(date, date2)
                print("different dates",sheet_name, row, order_id )
                mark_neutral(ws, row, worning_color.DATES_PROBLEM)

    wb.save(outputfile)

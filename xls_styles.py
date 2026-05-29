from openpyxl import load_workbook

def get_cell_formatting(file_path, sheet_name, cell_address):
    """
    Повертає інформацію про ячейку:
    - текст
    - закреслення
    - колір тексту
    - колір фону
    """

    # відкриваємо Excel
    workbook = load_workbook(file_path)

    # перевірка листа
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не знайдено")

    sheet = workbook[sheet_name]

    # отримуємо ячейку
    cell = sheet[cell_address]

    # значення
    value = cell.value

    # стиль шрифта
    font = cell.font

    # закреслення
    is_strikethrough = font.strike

    # колір тексту
    text_color = None

    if font.color is not None:
        if font.color.type == "rgb":
            text_color = font.color.rgb
        elif font.color.type == "theme":
            text_color = f"theme:{font.color.theme}"
        elif font.color.type == "indexed":
            text_color = f"indexed:{font.color.indexed}"

    # колір фону
    fill_color = None

    fill = cell.fill

    if fill.fill_type is not None:
        fg = fill.fgColor

        if fg.type == "rgb":
            fill_color = fg.rgb
        elif fg.type == "theme":
            fill_color = f"theme:{fg.theme}"
        elif fg.type == "indexed":
            fill_color = f"indexed:{fg.indexed}"

    return {
        "cell": cell_address,
        "value": value,
        "is_strikethrough": is_strikethrough,
        "text_color": text_color,
        "fill_color": fill_color,
    }
def print_cell_formatting(info):
    """
    Красивий друк інформації про ячейку
    """

    print(f"Ячейка: {info['cell']}")
    print(f"Текст: {info['value']}")
    print(f"Закреслення: {info['is_strikethrough']}")
    print(f"Колір тексту: {info['text_color']}")
    print(f"Колір фону: {info['fill_color']}")
def test_styles():
    file_path = "накази_втрати майна  А4007.xlsx"
    sheet_name = "Sheet1"

    # виклик для B2
    info = get_cell_formatting(
        file_path=file_path,
        sheet_name=sheet_name,
        cell_address="I5501"
    )

    print_cell_formatting(info)

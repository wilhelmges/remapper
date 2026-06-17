import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import sqlite3
from core import is_valid_date
from datetime import datetime

from core import headers, department_to_sheet, get_order_from_comment
from prepare_excel import prepare_excel
from styles import mark_neutral
from config import orders_network_url, sourcefile

# Відкрити базу (або створити, якщо її немає)
conn = sqlite3.connect("wasted.db")
cur = conn.cursor()

def find_root_order_info(ws, row):
    while True:
        row+=1
        changed = ws.cell(row=row, column=2).font.strike
        if changed:
            continue
        textdate = ws.cell(row=row, column=2).value
        operation = ws.cell(row=row, column=3).value
        order_id = get_order_from_comment(operation)
        if not is_valid_date(textdate) or order_id is None:
            return None, None
        date = datetime.date(textdate)
        return order_id, date

def extract():
    shutil.copy2(orders_network_url, sourcefile)
    prepare_excel()
    cur.execute("DELETE FROM wasted");  conn.commit()
    wb = load_workbook(sourcefile, data_only=True); ws:Worksheet = wb["Sheet1"]
    last_row =  ws.max_row #8000 #9475 6078 #
    row = 5995

    #print(ws.cell(row=6006, column=9).value); exit()

    for row in range(row, last_row):
        if ws.cell(row=row, column=2).value is None:
            continue

        textdate = ws.cell(row=row, column=2).value
        operation = str(ws.cell(row=row, column=3).value)
        order_id = get_order_from_comment(operation)

        if not is_valid_date(textdate) or order_id is None:
            print('cant get order or data', row, textdate, order_id, operation)
            continue

        values = [
        ws.cell(row=row, column=col).value
        for col in range(4, 25)
        ]

        w=[]
        blank = True
        for i, col in enumerate(range(4, 25)):
            #print(i, ws.cell(row=row, column=col).font.strike, headers[i], department_to_sheet[headers[i]], values[i])
            if ws.cell(row=row, column=col).font.strike or department_to_sheet[headers[i]]==None or values[i]==None:
                values[i]=None
            else:
                blank = False

        if blank:
            print('empty values ', row, values)
            mark_neutral(ws, row)

        changed = None; root_order_id = None; root_date = None
        date = datetime.date(textdate)
        changed = ws.cell(row=row, column=2).font.strike
        impacted = 1 if changed else 0
        if changed:
            root_order_id, root_date = find_root_order_info(ws, row)
            if root_order_id is None or root_date is None:
                print(row, "cant be processed")
                continue

        #print(row, order_id, date, changed, root_order_id, root_date, values )

        for i, value in enumerate(values):
            sheet_name = department_to_sheet[headers[i]]
            if value is not None:
                if sheet_name is not None:
                    pass#print('adding to sheet', sheet_name)
                else:
                    print('no sheet to add', headers[i])
                    print(row, order_id, date, changed, root_order_id, root_date, values)
                cur.execute(
                    "INSERT INTO wasted (order_id, order_date, root_order_id, root_date, department,sheet_name, money, row, impacted) "
                    "VALUES (?, ?, ?, ?, ?,?, ?, ?, ?)",
                    (order_id, date, root_order_id, root_date, headers[i],sheet_name, value, row, impacted)
                )
                if cur.lastrowid is None or cur.lastrowid==0:
                    print('couldnt insert for row ',row)

    conn.commit(); conn.close()
    wb.save(sourcefile)


if __name__=="__main__":
    extract()
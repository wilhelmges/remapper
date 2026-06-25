import hashlib
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime

from config import sourcefile
from utils.core import is_valid_order_row

def row_signature(ws, row=4):
    if not is_valid_order_row(ws, row):
        return None
    values = []
    for cell in ws[row]:
        value = (cell.value)
        if value is None:
            value = " ___ "
        elif isinstance(value, datetime):
            value = value.isoformat()
        values.append(str(value).strip())
    if ws.cell(row=row, column=2).font.strike:
        values.append('-')
    data = "\x1f".join(values)
    return hashlib.md5(data.encode("utf-8")).hexdigest()

def get_list_of_sigmatures_from_xlsx(filename):
    wb = load_workbook(filename, data_only=True); ws:Worksheet = wb["Sheet1"]
    last_row = ws.max_row
    row = 3

    rows=[]
    for row in range(row, last_row):
        sigmature = row_signature(ws, row)
        if sigmature:
            rows.append(sigmature)
    return rows

if __name__=="__main__":
    # wb = load_workbook(sourcefile, data_only=True); ws:Worksheet = wb["Sheet1"]
    rs = get_list_of_sigmatures_from_xlsx(sourcefile)
    print(len(rs))
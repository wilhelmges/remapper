from openpyxl import load_workbook
from utils.core import is_valid_date, sourcefile
import datetime
import re
from utils.row_marker import mark_row_wcolor,Warning_color

from utils.core import outputfile


def get_order(s= "№106 (ОД) 08.01.2026"):
    match = re.search(r"\d+", s)
    num = int(match.group()) if match else None
    return num

def validate_journal():
    year = 2026
    wb = load_workbook(outputfile, data_only=True)
    for ws in wb.worksheets:

        sheet_name = ws.title
        print(sheet_name)

        row = 8
        last_row = ws.max_row #14#
        for row in range(8, last_row):
            date = ws.cell(row=row, column=1).value
            if not is_valid_date(date):
                continue
            if type(date) == datetime.datetime:
                date = date.date()
            if type(date) == str:
                date = datetime.datetime.strptime(date, "%d.%m.%Y").date()
            order_id = get_order(ws.cell(row=row, column=2).value)
            d = (ws.cell(row=row, column=2).value.split()[-1])
            try:
                if len(d)==10:
                    date2 = datetime.datetime.strptime(d, "%d.%m.%Y").date()
                elif len(d)==8:
                    date2 = datetime.datetime.strptime(d, "%d.%m.%y").date()
                else:
                    date2 = None; print(" cant convert date: ", date, sheet_name, row, repr(d))
                    mark_row_wcolor(ws, row, Warning_color.DATES_PROBLEM)
            except Exception as e:
                print(" error while converting to date: ",date, sheet_name, row, repr(d))
                mark_row_wcolor(ws, row, Warning_color.DATES_PROBLEM)
                date2 = date

            if date!=date2:
                print(date, date2)
                print("different dates",sheet_name, row, order_id )
                mark_row_wcolor(ws, row, Warning_color.DATES_PROBLEM)

    wb.save(outputfile)

def validate_orders():
    wb = load_workbook(sourcefile, data_only=True)
    for ws in wb.worksheets:

        sheet_name = ws.title
        print(sheet_name)

if __name__ == "__main__":
    validate_journal()
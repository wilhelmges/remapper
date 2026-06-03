import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import sqlite3
import re
from core import is_valid_date
from datetime import datetime
from styles import mark_neutral
from openpyxl.styles import Font


from core import headers, department_to_sheet
import re

from core import get_order_from_comment, format_date_for_output

outputfile = "output.xlsx"

# Відкрити базу (або створити, якщо її немає)
conn = sqlite3.connect("wasted.db")
conn.row_factory = sqlite3.Row
cur = conn.cursor()

def find_start_blanc(ws):
    row = 8
    while True:
        if all(ws.cell(row=row + i, column=2).value is None for i in range(3)):
            return row+1
        row += 1

if __name__=="__main__":
    shutil.copy2("книга втрат електронний варіант.xlsx", outputfile)
    year = 2026
    wb = load_workbook(outputfile, data_only=True)
    for ws in wb.worksheets:
        cur.execute("""
            SELECT *
            FROM wasted
            WHERE 
            sheet_name= ?
            AND order_date >= ?
            AND order_date < ?
            AND filled=0 AND impacted=0
        """, (f"{ws.title}",f"{year}-01-01", f"{year + 1}-01-01"))

        rows = cur.fetchall()
        sheet_name = ws.title
        start_blanc = find_start_blanc(ws)
        #print(sheet_name, len(rows), start_blanc)

        i = start_blanc
        for row in rows:
            cell1 = ws.cell(row=i, column=1)
            cell1.value = str(format_date_for_output(row['order_date']));cell1.font=Font()
            cell2 = ws.cell(row=i, column=2)
            cell2.value="№" +str(row['order_id']) + " (ОД) "+format_date_for_output(row['order_date']);cell2.font=Font()
            cell3 = ws.cell(row=i, column=3)
            cell3.value = row['money']; cell3.font=Font()
            i+=1

            root_id = row['order_id']
            cur.execute("""
                        SELECT *
                        FROM wasted
                        WHERE
                        root_order_id= ?
                    """, (root_id,))

            nested_rows = cur.fetchall()
            jmax = len(nested_rows)
            print(sheet_name,root_id, jmax)

            for j,nested in enumerate(nested_rows):
                cell1 = ws.cell(row=i+j, column=2)
                cell1.value = "№" + str(nested['order_id']) + " (ОД) " + format_date_for_output(
                    nested['order_date']);cell1.font = Font(color="FF0000")
                cell2 = ws.cell(row=i+j, column=3)
                cell2.value = nested['money']; cell2.font = Font(color="FF0000")

            i+=jmax
            if jmax==0:
                i+=1

                # cur.execute(
                #     "UPDATE wasted SET filled = 1 WHERE id = ?",
                #     (order_id, sheet_name)
                # )
                # updated_rows = cur.rowcount
                # if updated_rows == 0:
                #     print("Запис не знайдено", order_id, date, sheet_name)
                # elif updated_rows == 1:
                #     pass #print("Оновлено 1 запис")


    conn.commit()
    conn.close()
    wb.save(outputfile)

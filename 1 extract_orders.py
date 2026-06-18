import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import sqlite3
from utils.core import is_valid_date
from datetime import datetime

from utils.core import headers, department_to_sheet, get_order_from_comment
from utils.prepare_excel import prepare_excel
from utils.row_marker import mark_row_wcolor, Warning_color
from config import orders_network_url, sourcefile

# Відкрити базу (або створити, якщо її немає)
conn = sqlite3.connect("wasted.db")
cur = conn.cursor()

def find_root_order_info(ws, row):
    while True:
        row+=1
        changed = ws.cell(row=row, column=2).font.strike
        if changed:
            continue
        textdate = ws.cell(row=row, column=2).value
        operation = ws.cell(row=row, column=3).value
        order_id = get_order_from_comment(operation)
        if not is_valid_date(textdate) or order_id is None:
            return None, None
        date = datetime.date(textdate)
        return order_id, date

def _extract():
    shutil.copy2(orders_network_url, sourcefile)
    prepare_excel(sourcefile)

    cur.execute("DELETE FROM wasted");  conn.commit()
    wb = load_workbook(sourcefile, data_only=True); ws:Worksheet = wb["Sheet1"]
    last_row =  ws.max_row #8000 #9475 6078 #
    row = 5995

    for row in range(row, last_row):
        #can date and order num be extracted
        if ws.cell(row=row, column=2).value is None:
            continue
        textdate = ws.cell(row=row, column=2).value
        operation = str(ws.cell(row=row, column=3).value)
        order_id = get_order_from_comment(operation)
        if not is_valid_date(textdate) or order_id is None:
            print('cant get order or data', row, textdate, order_id, operation)
            continue

        #checking wasted for services, if blanc - no real values
        values = [
        ws.cell(row=row, column=col).value
        for col in range(4, 25)
        ]
        w=[]
        blank = True
        for i, col in enumerate(range(4, 25)):
            if ws.cell(row=row, column=col).font.strike or department_to_sheet[headers[i]]==None or values[i]==None:
                values[i]=None
            else:
                blank = False
        if blank:
            #print('empty values ', row, values)
            mark_row_wcolor(ws, row, Warning_color.NEUTRAL)

        #if values were modified by another order, then finding this order
        changed = None; root_order_id = None; root_date = None
        date = datetime.date(textdate)
        changed = ws.cell(row=row, column=2).font.strike
        impacted = 1 if changed else 0
        if changed:
            root_order_id, root_date = find_root_order_info(ws, row)
            if root_order_id is None or root_date is None:
                print(row, "cant be processed")
                continue

        #adding wastes to sqlite
        for i, value in enumerate(values):
            sheet_name = department_to_sheet[headers[i]]
            if value is not None:
                if sheet_name is not None:
                    pass#print('adding to sheet', sheet_name)
                else:
                    print('no sheet to add', headers[i])
                    print(row, order_id, date, changed, root_order_id, root_date, values)
                cur.execute(
                    "INSERT INTO wasted (order_id, order_date, root_order_id, root_date, department,sheet_name, money, row, impacted) "
                    "VALUES (?, ?, ?, ?, ?,?, ?, ?, ?)",
                    (order_id, date, root_order_id, root_date, headers[i],sheet_name, value, row, impacted)
                )
                if cur.lastrowid is None or cur.lastrowid==0:
                    print('couldnt insert for row ',row)

    conn.commit(); conn.close()
    wb.save(sourcefile)

if __name__=="__main__":
    _extract()
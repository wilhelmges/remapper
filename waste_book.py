from pathlib import Path
from dataclasses import dataclass
from decimal import Decimal
import datetime
from types import SimpleNamespace


from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from utils.core import cell_to_sqlite_date, get_order_from_comment, cell_to_decimal, safe_decimal
from utils.row_marker import Warning_color, mark_row_wcolor

from config import wasted_backup as file_path, sample_xlsx, wasted_backup  # "data.xlsx"
sheets = ['РАО', 'ЗББ та Р', 'Зас ураж', 'НСО', 'БПЛА', 'ППО', 'ОВТ та МСП', 'РЕБ', 'Інж', 'РХБЗ', 'Реч', 'Звяз', 'Прод', 'ПММ', 'Мед', 'Авто', 'КЕС', 'Елек', 'Пожежна', 'Засоби розвідки', 'Гео', 'Метр']

# @dataclass()
# class Title_waste(init=False):
#     date: str
#     ordernum: int
#     updated: bool
#     title: str
#     titlesum: Decimal

def parse_sheet(ws: Worksheet):
    last_row =  ws.max_row # 8000 #9475 6078 #
    row = 7
    previous_order_num = None; set_total_sum = 0
    for row in range(row, last_row):
        date = cell_to_sqlite_date(ws.cell(row=row, column=2))
        ordernum = (get_order_from_comment(ws.cell(row=row, column=3).value))
        #print(row, date, ordernum, type(ordernum))
        #not data row
        if (date is None or ordernum is None) and previous_order_num is None:
            print(row, 'empty row')
            continue

        # new set started
        elif ((date and ordernum) and (previous_order_num != ordernum)) or ((date is None or ordernum is None) and previous_order_num is not None):
            if previous_order_num is not None:
                set_total_sum2 = ( cell_to_decimal(ws.cell(row=row-1, column=9)))
                if set_total_sum2 is not None and abs(abs(set_total_sum) - abs(set_total_sum2))>0.001:
                    mark_row_wcolor(ws, row-1, Warning_color.DIFF_SUMS)
                    print('does not equal for ', row-1, set_total_sum, set_total_sum2)
                else:
                    pass #print(set_total_sum, set_total_sum2)
            #print('------------------------------')
            #print('new set started with ordernum ', ordernum, type(ordernum))
            previous_order_num = ordernum
            set_total_sum = Decimal(0)
            changed = True if ws.cell(row=row, column=2).font.strike else False
            title = ws.cell(row=row, column=4).value
            title_total_sum = safe_decimal(ws.cell(row=row, column=8).value)
            set_total_sum += title_total_sum

            # print(row, date, ordernum, changed, title, title_total_sum)

        elif (date and ordernum) and (previous_order_num == ordernum and previous_order_num):
            # print('set continued', row, ordernum, type(ordernum))
            changed = True if ws.cell(row=row, column=2).font.strike else False
            title = ws.cell(row=row, column=4).value
            title_total_sum = safe_decimal(ws.cell(row=row, column=8).value)
            set_total_sum+=title_total_sum

            #print(row, date, ordernum, changed, title, title_total_sum)

            if title_total_sum is None:
                pass #print('file may be need to recalculate formulas'); exit(1)
                continue
        else:
            print(row, 'else')


def check_spoiled_orders(ws: Worksheet):
    all_present = True
    last_row = ws.max_row  # 8000 #9475 6078 #
    row = 12
    spoiled_rows = []
    for row in range(row, last_row):
        striked = True if (ws.cell(row=row, column=2).font.strike or ws.cell(row=row, column=9).font.strike) else False
        if not striked and isinstance(ws.cell(row=row, column=9).value, float):
            year = None
            for col in range(10, 15):
                if isinstance(ws.cell(row=row, column=col).value, float):
                    year = 2012 + col
                    #print(row, col, year)
                    break
            if year is None:
                print(' no year for ', row)
                all_present = False
                spoiled_rows.append(row)
            else:
                pass #ws.cell(row=row, column=1, value=year)

    if not all_present:
        ws.sheet_properties.tabColor = "FF0000"

def merged(ws: Worksheet):
    cell = "D6"
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            print("Головна комірка:", merged_range.start_cell.coordinate)

if __name__ == '__main__':
    wb = load_workbook(wasted_backup, data_only=True); ws:Worksheet = wb["ЗББ та Р"]
    #for ws in wb.worksheets:
    rows = check_spoiled_orders(ws)
    wb.save(wasted_backup); wb.close()




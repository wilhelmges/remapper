from openpyxl import load_workbook

# Шлях до Excel-файлу
from config import wasted_backup as file_path # "data.xlsx"
wb = load_workbook(file_path)

# Вивести список усіх листів
print("Листи у файлі:")
list = []

for i, sheet_name in enumerate(wb.sheetnames, start=1):
    list.append(sheet_name)

print(list)
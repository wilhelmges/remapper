from openpyxl import load_workbook
from config import wasted_backup
wb = load_workbook(wasted_backup)

sheet_name = "Count"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

wb.save(wasted_backup)
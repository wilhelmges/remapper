from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from config import sample_xlsx, wasted_backup

if __name__ == '__main__':
    wb = load_workbook(sample_xlsx, data_only=True);
    ws:Worksheet = wb["БПЛА"]
    value = ws["I8079"].value
    print(value)
    #for ws in wb.worksheets:
    ws.cell(row=1, column=1, value=42)
    wb.save(sample_xlsx); wb.close()

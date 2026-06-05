import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import sqlite3
import re
from core import is_valid_date
import datetime
import traceback


from core  import outputfile,headers, department_to_sheet
import re
from core import get_order_from_comment
from styles import mark_neutral, worning_color


def get_order(s= "№106 (ОД) 08.01.2026"):
    match = re.search(r"\d+", s)
    num = int(match.group()) if match else None
    return num



# Відкрити базу (або створити, якщо її немає)
conn = sqlite3.connect("wasted.db")
cur = conn.cursor()

if __name__=="__main__":
    year = 2026
    wb = load_workbook(outputfile, data_only=True)
    for ws in wb.worksheets:
        cur.execute("""
            SELECT COUNT(*)
            FROM wasted
            WHERE 
            sheet_name= ?
            AND order_date >= ?
            AND order_date < ?
        """, (f"{ws.title}",f"{year}-01-01", f"{year + 1}-01-01"))

        count = cur.fetchone()[0]
        if count<1:
            continue
        sheet_name = ws.title
        print(sheet_name, count)

        row = 8
        last_row = ws.max_row #14#
        for row in range(8, last_row):
            date = ws.cell(row=row, column=1).value
            if not is_valid_date(date):
                continue
            if type(date)==datetime.datetime:
                date = date.date()
            if type(date) == str:
                date = datetime.datetime.strptime(date, "%d.%m.%Y").date()
            money = ws.cell(row=row, column=3).value
            #print(type(date), sheet_name, row, date)
            order_id = get_order(ws.cell(row=row, column=2).value)

            cur.execute(
                "UPDATE wasted SET filled = 1 WHERE order_id = ? AND order_date=?  AND sheet_name = ? AND money=?",
                (order_id, date, sheet_name, money)
            )
            updated_rows = cur.rowcount
            if updated_rows == 0:
                print("Запис не знайдено", order_id, date, sheet_name)
                mark_neutral(ws, row, worning_color.ORDER_NOT_FOUND)
            elif updated_rows == 1:
                pass #print("Оновлено 1 запис")
            else:
                print("Оновлено більше одного записа", order_id, date, sheet_name)
                mark_neutral(ws, row, worning_color.MULTIPLE_RECORD)

    conn.commit()
    conn.close()
    wb.save(outputfile)
